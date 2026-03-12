from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl import load_workbook
import io, os, zipfile, shutil, tempfile

app = Flask(__name__)
CORS(app)

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template.xlsx')

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

@app.route('/generate', methods=['POST'])
def generate():
    data = request.json

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb['履歴書']

    def sc(addr, val):
        ws[addr] = str(val) if val else ''

    # 記入日
    from datetime import date
    today = date.today()
    sc('E3', f'　　{today.year}年{today.month}月{today.day}日現在')

    # 基本情報
    sc('C6', (data.get('lastKana','') + '　' + data.get('firstKana','')).strip())
    sc('C9', (data.get('lastName','') + '　' + data.get('firstName','')).strip())

    # 性別
    sc('F15', data.get('gender',''))

    # 生年月日
    by = data.get('birthYear','')
    bm = data.get('birthMonth','')
    bd = data.get('birthDay','')
    if by:
        try:
            birth = date(int(by), int(bm), int(bd))
            today2 = date.today()
            age = today2.year - birth.year - ((today2.month, today2.day) < (birth.month, birth.day))
            sc('B14', f"{by}年　{str(bm).zfill(2)}月　{str(bd).zfill(2)}日生　（満　{age}　歳）")
        except:
            sc('B14', f"{by}年　{bm}月　{bd}日生")

    # 住所
    sc('C16', data.get('addrKana',''))
    sc('I16', ' ' + data.get('phone',''))
    sc('C19', '〒' + data.get('zipCode',''))
    sc('C21', data.get('address',''))
    sc('H21', ' ' + data.get('email',''))

    # 学歴・職歴
    EDU_L = [35,38,41,44,47,50,53,56,59,62,65,68,71,74,77,80]
    EDU_R = [5,8,11,14,16,19]

    ent = []
    education = data.get('education', [])
    jobs = data.get('jobs', [])

    if education:
        ent.append({'yr':'','mo':'','txt':'学　歴'})
        for e in education:
            if e.get('enterYear'):
                ent.append({'yr':e['enterYear'],'mo':e.get('enterMonth',''),'txt':f"{e['name']}　入学"})
            if e.get('exitYear'):
                ent.append({'yr':e['exitYear'],'mo':e.get('exitMonth',''),'txt':f"{e['name']}　卒業"})

    if jobs:
        ent.append({'yr':'','mo':'','txt':'職　歴'})
        for j in jobs:
            if j.get('enterYear'):
                ent.append({'yr':j['enterYear'],'mo':j.get('enterMonth',''),'txt':f"{j['name']}　入社（{j.get('type','')}）"})
            if j.get('exitYear'):
                ent.append({'yr':j['exitYear'],'mo':j.get('exitMonth',''),'txt':f"{j['name']}　退社"})
            elif j.get('enterYear'):
                ent.append({'yr':'','mo':'','txt':'現在に至る'})

    ent.append({'yr':'','mo':'','txt':'以　上'})

    lft = ent[:16]
    rgt = ent[16:22]

    for i, r in enumerate(EDU_L):
        if i < len(lft):
            sc(f'B{r}', lft[i]['yr'])
            sc(f'C{r}', lft[i]['mo'])
            sc(f'D{r}', lft[i]['txt'])
        else:
            sc(f'B{r}', '')
            sc(f'C{r}', '')
            sc(f'D{r}', '')

    for i, r in enumerate(EDU_R):
        if i < len(rgt):
            sc(f'L{r}', rgt[i]['yr'])
            sc(f'M{r}', rgt[i]['mo'])
            sc(f'N{r}', rgt[i]['txt'])
        else:
            sc(f'L{r}', '')
            sc(f'M{r}', '')
            sc(f'N{r}', '')

    # ヘッダー行を復元（書き込みで上書きされた場合の対策）
    from openpyxl.styles import Font, Border, Side
    hdr_font = Font(name='ＭＳ Ｐ明朝', size=10)

    def set_hdr(addr, val, double_bottom=False):
        ws[addr] = val
        ws[addr].font = hdr_font
        if double_bottom:
            ws[addr].border = Border(
                top=Side(border_style='thin'),
                bottom=Side(border_style='double'),
                left=Side(border_style='thin'),
                right=Side(border_style='thin')
            )

    # 左ページ（B35/C35/D35の下線を二重線）
    set_hdr('B35', '年', double_bottom=True)
    set_hdr('C35', '月', double_bottom=True)
    set_hdr('D35', '学  歴 ・ 職  歴 （各別にまとめて書く）', double_bottom=True)
    # 右ページ学歴職歴（L2/M2/N2の下線を二重線）
    set_hdr('L2', '年', double_bottom=True)
    set_hdr('M2', '月', double_bottom=True)
    set_hdr('N2', '学  歴 ・ 職  歴 （各別にまとめて書く）', double_bottom=True)
    # 右ページ資格
    set_hdr('L22', '年')
    set_hdr('M22', '月')
    set_hdr('N22', '資  格 ・ 免  許')

    # 資格
    LIC_R = [25,28,31,34,37,40]
    licenses = data.get('licenses', [])
    for i, r in enumerate(LIC_R):
        if i < len(licenses):
            sc(f'L{r}', licenses[i].get('year',''))
            sc(f'M{r}', licenses[i].get('month',''))
            sc(f'N{r}', licenses[i].get('name',''))
        else:
            sc(f'L{r}', '')
            sc(f'M{r}', '')
            sc(f'N{r}', '')

    # 志望動機
    sc('L47', data.get('prText',''))

    # 本人希望欄
    hp = []
    if data.get('currentSalary'): hp.append(f"現在年収：{data['currentSalary']}万円")
    if data.get('desiredSalary'): hp.append(f"希望年収：{data['desiredSalary']}万円")
    if data.get('pcSkills'):      hp.append(f"PCスキル：{data['pcSkills']}")
    if data.get('hopeText'):      hp.append(data['hopeText'])
    HOPE_ROWS = [71, 74, 77, 80, 83]
    for i, r in enumerate(HOPE_ROWS):
        sc(f'L{r}', hp[i] if i < len(hp) else '')

    # openpyxlで一時保存
    tmp_out = io.BytesIO()
    wb.save(tmp_out)
    tmp_out.seek(0)

    # テンプレートのdrawing1.xmlを出力ファイルに移植
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        tmp_path = f.name
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        out_path = f.name

    # テンプレートからdrawingファイルを取得
    with zipfile.ZipFile(TEMPLATE_PATH, 'r') as tmpl_zip:
        drawing_files = [n for n in tmpl_zip.namelist() if 'drawing' in n]

    # 出力ファイルにdrawingを追加
    with zipfile.ZipFile(io.BytesIO(tmp_out.getvalue()), 'r') as out_zip:
        with zipfile.ZipFile(TEMPLATE_PATH, 'r') as tmpl_zip:
            with zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as new_zip:
                # 出力ファイルの全ファイルをコピー（drawing関連は除く）
                for item in out_zip.namelist():
                    if 'drawing' not in item and '_rels/sheet1' not in item:
                        new_zip.writestr(item, out_zip.read(item))
                # テンプレートのdrawingファイルとsheet1のrelsをコピー
                for item in tmpl_zip.namelist():
                    if 'drawing' in item or '_rels/sheet1' in item:
                        new_zip.writestr(item, tmpl_zip.read(item))

    with open(out_path, 'rb') as f:
        output = io.BytesIO(f.read())
    os.unlink(out_path)
    output.seek(0)

    name = (data.get('lastName','') + data.get('firstName','')).strip() or '応募者'
    filename = f"履歴書_{name}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
